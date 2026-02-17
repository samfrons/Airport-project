#!/usr/bin/env python3
"""
JPX Dashboard — Noise Complaint Data Loader
============================================

Loads complaint data from PlaneNoise CSV/Excel exports into the database.
Performs:
  - PII stripping (removes names, phone, email, house numbers)
  - Date/time parsing and UTC conversion
  - Street-level geocoding with caching
  - Flight correlation matching
  - Daily summary and hotspot aggregation

Usage:
    python scripts/load_complaints.py complaints.csv
    python scripts/load_complaints.py complaints.xlsx --sheet "Complaints"
    python scripts/load_complaints.py complaints.csv --skip-geocoding
    python scripts/load_complaints.py complaints.csv --dry-run

Expected CSV columns (PlaneNoise export format):
    - Date of Noise Event
    - Time of Noise Event
    - Street Address (or Address)
    - City/Town (or Municipality)
    - Zip Code
    - Airport
    - Complaint Type(s) (comma-separated)
    - Aircraft Type
    - Aircraft Description
    - Direction
    - Comments
    - Submission Date
"""

import sys
import re
import csv
import json
import logging
import click
import time
from pathlib import Path
from datetime import datetime, timezone
from typing import Optional
from zoneinfo import ZoneInfo

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.db.database import (
    get_connection, init_db, insert_complaint,
    update_complaint_daily_summary, update_complaint_hotspots,
    get_geocode_cache, cache_geocode, get_flights_in_window
)
from src.analysis.classify import is_curfew_hour, is_weekend

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

ET = ZoneInfo("America/New_York")

# ── Column Name Mappings ─────────────────────────────────────────────────────
# PlaneNoise exports may use various column names. Map them to our schema.

COLUMN_MAPPINGS = {
    "event_date": ["Date of Noise Event", "Event Date", "Date", "date"],
    "event_time": ["Time of Noise Event", "Event Time", "Time", "time"],
    "address": ["Street Address", "Address", "Street", "address", "street_address"],
    "municipality": ["City/Town", "Municipality", "City", "Town", "city", "municipality"],
    "zip_code": ["Zip Code", "ZIP", "Zip", "zip_code", "zipcode"],
    "airport": ["Airport", "airport"],
    "complaint_types": ["Complaint Type(s)", "Complaint Types", "Complaint Type", "Type of Complaint", "complaint_types"],
    "aircraft_type": ["Aircraft Type", "Type of Aircraft", "aircraft_type"],
    "aircraft_description": ["Aircraft Description", "Description of Aircraft", "aircraft_description"],
    "flight_direction": ["Direction", "Flight Direction", "direction"],
    "comments": ["Comments", "Additional Comments", "Notes", "comments"],
    "submission_date": ["Submission Date", "Date Submitted", "Filed Date", "submission_date"],
}


def find_column(headers: list, field: str) -> Optional[str]:
    """Find the actual column name for a field from possible mappings."""
    possible = COLUMN_MAPPINGS.get(field, [field])
    for col in possible:
        if col in headers:
            return col
    return None


# ── PII Stripping ────────────────────────────────────────────────────────────

def strip_house_number(address: str) -> str:
    """
    Remove house number from street address, keeping only street name.

    Examples:
        "123 Main Street" -> "Main Street"
        "45A Ocean Ave" -> "Ocean Ave"
        "Main Street" -> "Main Street"
    """
    if not address:
        return ""

    # Pattern matches: leading numbers, optional letter suffix, and spaces
    # e.g., "123 ", "45A ", "12-14 ", "100B "
    pattern = r"^\d+[-\d]*[A-Za-z]?\s+"
    return re.sub(pattern, "", address.strip())


def normalize_municipality(municipality: str) -> str:
    """Normalize municipality names for consistency."""
    if not municipality:
        return ""

    # Common variations
    replacements = {
        "East Hampton": "East Hampton",
        "Easthampton": "East Hampton",
        "Sag Harbor": "Sag Harbor",
        "Sagaponack": "Sagaponack",
        "Wainscott": "Wainscott",
        "Bridgehampton": "Bridgehampton",
        "Amagansett": "Amagansett",
        "Springs": "Springs",
        "Montauk": "Montauk",
        "Noyac": "Noyac",
        "North Haven": "North Haven",
        "Shelter Island": "Shelter Island",
        "Southampton": "Southampton",
        "Water Mill": "Water Mill",
        "Watermill": "Water Mill",
    }

    normalized = municipality.strip().title()
    return replacements.get(normalized, normalized)


# ── Time Parsing ─────────────────────────────────────────────────────────────

def parse_event_datetime(date_str: str, time_str: str) -> tuple:
    """
    Parse event date and time strings into components.

    Returns:
        (event_date, event_time, event_datetime_utc, event_hour_et)
    """
    event_date = None
    event_time = None
    event_datetime_utc = None
    event_hour_et = None

    # Parse date
    if date_str:
        for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d-%m-%Y"]:
            try:
                dt = datetime.strptime(date_str.strip(), fmt)
                event_date = dt.strftime("%Y-%m-%d")
                break
            except ValueError:
                continue

    # Parse time
    if time_str and event_date:
        time_str = time_str.strip().upper()
        for fmt in ["%H:%M", "%I:%M %p", "%I:%M%p", "%H:%M:%S"]:
            try:
                # Handle "AM"/"PM" without space
                normalized = re.sub(r"(\d)(AM|PM)", r"\1 \2", time_str)
                tm = datetime.strptime(normalized, fmt)
                event_time = tm.strftime("%H:%M")
                event_hour_et = tm.hour

                # Combine with date for UTC conversion
                full_dt = datetime.strptime(f"{event_date} {event_time}", "%Y-%m-%d %H:%M")
                full_dt = full_dt.replace(tzinfo=ET)
                event_datetime_utc = full_dt.astimezone(timezone.utc).isoformat()
                break
            except ValueError:
                continue

    return event_date, event_time, event_datetime_utc, event_hour_et


def parse_submission_date(date_str: str) -> Optional[str]:
    """Parse submission date string."""
    if not date_str:
        return None

    for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d-%m-%Y"]:
        try:
            dt = datetime.strptime(date_str.strip(), fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


# ── Geocoding ────────────────────────────────────────────────────────────────

def geocode_address(street: str, municipality: str, zip_code: str,
                    conn, use_cache: bool = True) -> tuple:
    """
    Geocode a street address to lat/lng.

    Uses cache first, then falls back to US Census Geocoder API.

    Returns:
        (latitude, longitude) or (None, None) if geocoding fails
    """
    if not street or not municipality:
        return None, None

    # Check cache
    if use_cache:
        cached = get_geocode_cache(conn, street, municipality)
        if cached:
            return cached["latitude"], cached["longitude"]

    # Try US Census Geocoder (free, no API key needed)
    try:
        import urllib.request
        import urllib.parse

        # Build address string
        address = f"{street}, {municipality}, NY"
        if zip_code:
            address += f" {zip_code}"

        # Census geocoder endpoint
        url = "https://geocoding.geo.census.gov/geocoder/locations/onelineaddress"
        params = urllib.parse.urlencode({
            "address": address,
            "benchmark": "Public_AR_Current",
            "format": "json"
        })

        req = urllib.request.Request(
            f"{url}?{params}",
            headers={"User-Agent": "JPXDashboard/1.0"}
        )

        with urllib.request.urlopen(req, timeout=10) as response:
            data = json.loads(response.read().decode())

        matches = data.get("result", {}).get("addressMatches", [])
        if matches:
            coords = matches[0]["coordinates"]
            lat = coords["y"]
            lng = coords["x"]

            # Cache the result
            cache_geocode(conn, street, municipality, zip_code or "", lat, lng, "census")

            return lat, lng

    except Exception as e:
        log.debug(f"Geocoding failed for '{street}, {municipality}': {e}")

    # Fallback: return municipality centroid if known
    municipality_centroids = {
        "Wainscott": (40.9381, -72.2401),
        "Bridgehampton": (40.9374, -72.3006),
        "Sag Harbor": (40.9976, -72.2940),
        "East Hampton": (40.9634, -72.1848),
        "Amagansett": (40.9726, -72.1395),
        "Montauk": (41.0359, -71.9545),
        "Springs": (40.9476, -72.1560),
        "Noyac": (40.9937, -72.3410),
        "Southampton": (40.8843, -72.3893),
        "North Haven": (41.0062, -72.3065),
        "Sagaponack": (40.9245, -72.2691),
        "Water Mill": (40.9148, -72.3456),
    }

    if municipality in municipality_centroids:
        lat, lng = municipality_centroids[municipality]
        cache_geocode(conn, street, municipality, zip_code or "", lat, lng, "centroid")
        return lat, lng

    return None, None


# ── Flight Correlation ───────────────────────────────────────────────────────

def match_complaint_to_flight(complaint: dict, conn) -> tuple:
    """
    Match a complaint to a flight based on time window and aircraft type.

    Returns:
        (matched_flight_id, matched_confidence, matched_registration, matched_operator)
    """
    datetime_utc = complaint.get("event_datetime_utc")
    if not datetime_utc:
        return None, "unmatched", None, None

    # Get flights within ±15 minutes
    candidates = get_flights_in_window(conn, datetime_utc, minutes=15)

    if not candidates:
        return None, "unmatched", None, None

    # Score each candidate
    complaint_aircraft = (complaint.get("aircraft_type") or "").lower()
    complaint_direction = (complaint.get("flight_direction") or "").lower()

    best_match = None
    best_score = 0

    for flight in candidates:
        score = 0

        # Time proximity (higher score for closer match)
        # Candidates are already sorted by time proximity
        score += 1

        # Aircraft type match
        flight_category = (flight.get("aircraft_category") or "").lower()
        if complaint_aircraft == "helicopter" and flight_category == "helicopter":
            score += 3
        elif complaint_aircraft == "jet" and flight_category == "jet":
            score += 3
        elif complaint_aircraft == "prop" and flight_category == "fixed_wing":
            score += 2

        # Direction match
        flight_direction = (flight.get("direction") or "").lower()
        if complaint_direction == flight_direction:
            score += 1

        # Solo flight bonus
        if len(candidates) == 1:
            score += 2

        if score > best_score:
            best_score = score
            best_match = flight

    if not best_match:
        return None, "unmatched", None, None

    # Determine confidence
    if best_score >= 5:
        confidence = "high"
    elif best_score >= 3:
        confidence = "medium"
    elif best_score >= 1:
        confidence = "low"
    else:
        confidence = "unmatched"

    return (
        best_match.get("fa_flight_id"),
        confidence,
        best_match.get("registration"),
        best_match.get("operator")
    )


# ── Main Loader ──────────────────────────────────────────────────────────────

def load_csv(filepath: Path, sheet: str = None) -> list:
    """Load complaints from CSV or Excel file."""
    rows = []

    suffix = filepath.suffix.lower()

    if suffix in [".xlsx", ".xls"]:
        try:
            import openpyxl
        except ImportError:
            log.error("openpyxl required for Excel files. Run: pip install openpyxl")
            sys.exit(1)

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))

    else:  # CSV
        with open(filepath, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)

    return rows


def process_row(row: dict, headers: list, conn, skip_geocoding: bool = False,
                skip_matching: bool = False) -> dict:
    """Process a single complaint row into a database-ready dict."""

    # Extract fields using column mappings
    def get(field):
        col = find_column(headers, field)
        return row.get(col, "") if col else ""

    # Parse date/time
    event_date, event_time, event_datetime_utc, event_hour_et = parse_event_datetime(
        get("event_date"), get("event_time")
    )

    if not event_date:
        return None  # Skip rows without valid date

    # Strip PII from address
    raw_address = get("address")
    street_name = strip_house_number(raw_address)

    # Normalize municipality
    municipality = normalize_municipality(get("municipality"))

    # Geocode
    latitude, longitude = None, None
    if not skip_geocoding and street_name and municipality:
        latitude, longitude = geocode_address(
            street_name, municipality, get("zip_code"), conn
        )

    # Build complaint record
    complaint = {
        "source_id": row.get("ID") or row.get("id") or None,
        "event_date": event_date,
        "event_time": event_time,
        "event_datetime_utc": event_datetime_utc,
        "event_hour_et": event_hour_et,
        "is_curfew_period": 1 if event_hour_et is not None and is_curfew_hour(event_hour_et) else 0,
        "is_weekend": 1 if event_date and is_weekend(event_date) else 0,
        "street_name": street_name,
        "municipality": municipality,
        "zip_code": get("zip_code"),
        "latitude": latitude,
        "longitude": longitude,
        "airport": get("airport") or "JPX",
        "complaint_types": get("complaint_types"),
        "aircraft_type": get("aircraft_type"),
        "aircraft_description": get("aircraft_description"),
        "flight_direction": get("flight_direction"),
        "comments": get("comments"),
        "submission_date": parse_submission_date(get("submission_date")),
    }

    # Flight correlation
    if not skip_matching and event_datetime_utc:
        (
            complaint["matched_flight_id"],
            complaint["matched_confidence"],
            complaint["matched_registration"],
            complaint["matched_operator"]
        ) = match_complaint_to_flight(complaint, conn)
    else:
        complaint["matched_flight_id"] = None
        complaint["matched_confidence"] = "unmatched"
        complaint["matched_registration"] = None
        complaint["matched_operator"] = None

    return complaint


@click.command()
@click.argument("filepath", type=click.Path(exists=True))
@click.option("--sheet", default=None, help="Excel sheet name (if xlsx)")
@click.option("--skip-geocoding", is_flag=True, help="Skip address geocoding")
@click.option("--skip-matching", is_flag=True, help="Skip flight correlation matching")
@click.option("--dry-run", is_flag=True, help="Process but don't insert into database")
@click.option("--rate-limit", default=1.0, type=float, help="Seconds between geocode requests")
def main(filepath: str, sheet: str, skip_geocoding: bool, skip_matching: bool,
         dry_run: bool, rate_limit: float):
    """Load noise complaints from CSV/Excel into the JPX Dashboard database."""

    filepath = Path(filepath)
    log.info(f"Loading complaints from {filepath}")

    # Ensure database exists
    db_path = Path(__file__).parent.parent / "data" / "jpx_flights.db"
    if not db_path.exists():
        log.info("Database not found — initializing...")
        db_path.parent.mkdir(exist_ok=True)
        init_db(str(db_path))

    conn = get_connection(str(db_path))

    # Load file
    rows = load_csv(filepath, sheet)
    if not rows:
        log.error("No data found in file")
        return

    headers = list(rows[0].keys())
    log.info(f"Found {len(rows)} rows with columns: {', '.join(headers[:5])}...")

    print(f"\n{'═' * 60}")
    print(f"  JPX Dashboard — Complaint Loader")
    print(f"  File: {filepath.name}")
    print(f"  Rows: {len(rows)}")
    print(f"  Geocoding: {'Skipped' if skip_geocoding else 'Enabled'}")
    print(f"  Flight matching: {'Skipped' if skip_matching else 'Enabled'}")
    print(f"  Dry run: {'Yes' if dry_run else 'No'}")
    print(f"{'═' * 60}\n")

    inserted = 0
    skipped = 0
    geocoded = 0
    matched = 0
    dates_affected = set()
    last_geocode = 0

    for i, row in enumerate(rows, 1):
        complaint = process_row(
            row, headers, conn,
            skip_geocoding=skip_geocoding,
            skip_matching=skip_matching
        )

        if not complaint:
            skipped += 1
            continue

        if complaint.get("latitude"):
            geocoded += 1

        if complaint.get("matched_confidence") in ("high", "medium", "low"):
            matched += 1

        if not dry_run:
            if insert_complaint(conn, complaint):
                inserted += 1
                if complaint.get("event_date"):
                    dates_affected.add(complaint["event_date"])
            else:
                skipped += 1
        else:
            inserted += 1  # Count as would-be-inserted for dry run

        # Rate limit geocoding
        if not skip_geocoding and complaint.get("street_name"):
            elapsed = time.time() - last_geocode
            if elapsed < rate_limit:
                time.sleep(rate_limit - elapsed)
            last_geocode = time.time()

        # Progress
        if i % 100 == 0:
            log.info(f"  Processed {i}/{len(rows)} rows...")

    # Update aggregation tables
    if not dry_run and dates_affected:
        log.info("Updating daily summaries...")
        for date in dates_affected:
            update_complaint_daily_summary(conn, date)

        log.info("Updating hotspots...")
        update_complaint_hotspots(conn)

    conn.close()

    # Summary
    print(f"\n{'─' * 60}")
    print(f"  Results:")
    print(f"  Processed:      {len(rows)}")
    print(f"  Inserted:       {inserted}")
    print(f"  Skipped:        {skipped}")
    print(f"  Geocoded:       {geocoded}")
    print(f"  Flight matches: {matched}")
    print(f"  Dates affected: {len(dates_affected)}")
    print(f"{'─' * 60}\n")


if __name__ == "__main__":
    main()
